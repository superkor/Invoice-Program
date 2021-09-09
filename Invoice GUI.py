from appJar import gui
import openpyxl, datetime, os, shutil

class InvalidFileError(Exception):
    pass

class InvalidFileNameError(Exception):
    pass

class FileAlreadyExistsError(Exception):
    pass

row = ["9", "11", "13", "15", "17"]
rowIndex = 0
col = ["B", "C", "D", "E", "F", "G", "H"]
colIndex = -1
day = 2
path = ""

leapyear = 0
year = float(datetime.datetime.today().year)

if (year//4).is_integer():
    if isinstance(year//100, float):
        if (year//400).is_integer():
            leapyear = 1
else:
    leapyear = 0

month = datetime.datetime.today().month
og = os.getcwd()
original = os.path.join(og+"\Invoices\Invoice template.xlsx")
wb = openpyxl.load_workbook(original)
sheet = wb["Timesheet"]
icon = os.path.join(og+"\MSC logo.ico")
os.chdir("Invoices")

created = -1

def fileSend(btn):
    app.showSubWindow("Use File")
    
def createFile(btn):
    app.showSubWindow("Create File")
    
def fileUse(btn):
    global wb, sheet, created, path
    try:
        if app.getEntry("Invoice File") == ((og+"\\Invoices\\Invoice template.xlsx").replace("\\", "/")):
            raise InvalidFileError
        else:
            wb = openpyxl.load_workbook((app.getEntry("Invoice File")))
            path = (app.getEntry("Invoice File"))
            sheet = wb["Timesheet"]
            created = 0
            invoice(btn)
            app.hideSubWindow("Use File")
            app.clearAllEntries()
    except openpyxl.utils.exceptions.InvalidFileException:
        app.errorBox("Invalid File", "You must enter a *.xlsx file for use.")
        app.showSubWindow("Use File")
        app.clearAllEntries()
    except InvalidFileError:
        app.errorBox("Invalid File", "You can not use the Invoice template.")
        app.showSubWindow("Use File")
        app.clearAllEntries()
    except openpyxl.utils.exceptions.ReadOnlyWorkbookException:
        app.errorBox("Error Opening File", "The speadsheet is set to read-only. Try closing the speadsheet or changing the settings and try again.")
        app.showSubWindow("Use File")
        app.clearAllEntries()
    except FileNotFoundError:
            app.errorBox("File Error", "File has either moved or does not exist at the path given. Please try again.")
            app.showSubWindow("Move")
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Use File")
    
def checkName(btn):
    global wb, sheet, created
    try:
        for i in os.listdir():
            if i == (app.getEntry("New File Name")+".xlsx"):
                raise FileAlreadyExistsError
        if app.getEntry("New File Name") == "":
            raise InvalidFileNameError
        if app.getEntry("New File Name") == "Invoice template.xlsx":
            raise InvalidFileError
        
        else:
            app.hideSubWindow("Create File")
            shutil.copy(original, og+"\\Invoices\\"+app.getEntry("New File Name")+".xlsx")
            wb = openpyxl.load_workbook(og+"\\Invoices\\"+app.getEntry("New File Name")+".xlsx")
            sheet = wb["Timesheet"]
            created = 1
            invoice(btn)
    except InvalidFileError:
            app.errorBox("Invalid File Name", "You must enter a file name.")
            app.showSubWindow("Create File")
            app.clearAllEntries()
    except InvalidFileError:
        app.errorBox("Invalid File Name", "You must enter another file name.")
        app.showSubWindow("Create File")
        app.clearAllEntries()
    except FileAlreadyExistsError:
        app.errorBox("Error Creating File", "File Name Already Exists. Please choose another file.")
        app.showSubWindow("Create File")
        app.clearAllEntries()
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Create File")
        

#Update existing info onto label
def invoice(btn):
    app.showSubWindow("Invoice")
    app.setLabel("Name Value", "Name: "+ str(sheet["J5"].value))
    app.setLabel("Wage Value", "Wage: "+ str(sheet["J7"].value))
    app.setLabel("PCS Value", "PCS: "+ str(sheet["C20"].value))
    app.setLabel("CS Value", "CS: "+ str(sheet["C21"].value))
    app.setLabel("IN Value", "IN: "+ str(sheet["C22"].value))
    app.setLabel("PEP Value", "PEP: "+ str(sheet["C23"].value))
    app.setLabel("AD Value", "AD: "+ str(sheet["C24"].value))
    app.setLabel("PW Value", "PW: "+ str(sheet["C25"].value))
    app.setLabel("ST Value", "ST: "+ str(sheet["C26"].value))
    app.setLabel("Off-ice Value", "Off-ice: "+ str(sheet["C27"].value))
    app.setLabel("Off-ice Min Value", "Off-ice Min: "+ str(sheet["D27"].value))
    app.setLabel("Other Value", "Other: "+ str(sheet["C28"].value))
    app.setLabel("Other Min Value", "Other Mins: "+ str(sheet["D28"].value))
    app.setLabel("Notes 1 Value", "Note 1: "+str(sheet["I20"].value))
    app.setLabel("Notes 2 Value", "Note 2: "+str(sheet["I21"].value))
    app.setLabel("Notes 3 Value", "Note 3: "+str(sheet["I22"].value))
    app.setLabel("Notes 4 Value", "Note 4: "+str(sheet["I23"].value))
    app.setLabel("Notes 5 Value", "Note 5: "+str(sheet["I24"].value))

def abort(btn):
    app.stop()

def cancelOpen(btn):
    app.hideAllSubWindows()
    app.clearAllEntries()

def about(btn):
    app.showSubWindow("About Me!")

def help(btn):
    app.warningBox("HELP!", "????????")

#Invoice info
def accept(btn):
    global sheet
    sheet = wb["Timesheet"]
    try:
        if app.getEntry("Name Input") is not None:
            sheet["J5"] = str(app.getEntry("Name Input"))
        if app.getEntry("Wage Input") is not None:
            sheet["J7"] = float(app.getEntry("Wage Input"))

        if app.getEntry("PCS") is not None:
            if sheet["C20"].value is not None:
                sheet["C20"] = int(app.getEntry("PCS")) + sheet["C20"].value
            else:
                sheet["C20"] = int(app.getEntry("PCS"))
        
        if app.getEntry("CS") is not None:
            if sheet["C21"].value is not None:
                sheet["C21"] = int(app.getEntry("CS")) + sheet["C21"].value
            else:
                sheet["C21"] = int(app.getEntry("CS"))
        
        if app.getEntry("IN") is not None:
            if sheet["C22"].value is not None:
                sheet["C22"] = int(app.getEntry("IN")) + sheet["C22"].value
            else:
                sheet["C22"] = int(app.getEntry("IN"))
        
        if app.getEntry("PEP") is not None:
            if sheet["C23"].value is not None:
                sheet["C23"] = int(app.getEntry("PEP")) + sheet["C23"].value
            else:
                sheet["C23"] = int(app.getEntry("PEP"))

        if app.getEntry("AD") is not None:
            if sheet["C24"].value is not None:
                sheet["C24"] = int(app.getEntry("AD")) + sheet["C24"].value
            else:
                sheet["C24"] = int(app.getEntry("AD"))
        
        if app.getEntry("PW") is not None:
            if sheet["C25"].value is not None:
                sheet["C25"] = int(app.getEntry("PW")) +sheet["C25"].value
            else:
                sheet["C25"] = int(app.getEntry("PW"))
        
        if app.getEntry("ST") is not None:
            if sheet["C26"].value is not None:
                sheet["C26"] = int(app.getEntry("ST")) + sheet["C26"].value
            else:
                sheet["C26"] = int(app.getEntry("ST"))

        if app.getEntry("Off Ice") is not None:
            if sheet["C27"].value is not None:
                sheet["C27"] = int(app.getEntry("Off Ice")) + sheet["C27"].value
            else:
                sheet["C27"] = int(app.getEntry("Off Ice"))
        
        if app.getEntry("Off Ice Mins") is not None:
            if sheet["D27"].value is not None:
                sheet["D27"] = int(app.getEntry("Off Ice Mins")) + sheet["D27"].value
            else:
                sheet["D27"] = int(app.getEntry("Off Ice Mins"))

        if app.getEntry("Other") is not None:
            if sheet["C28"].value is not None:
                sheet["C28"] = int(app.getEntry("Other")) + sheet["C28"].value
            else:
                sheet["C28"] = int(app.getEntry("Other"))

        if (app.getEntry("Other Mins") is not None) or (sheet["D28"].value is not None):
            if sheet["D28"].value is not None:
                sheet["D28"] = int(app.getEntry("Other Mins")) + sheet["D28"].value
            else:
                sheet["D28"] = int(app.getEntry("Other Mins"))
            
            sheet["E28"] = sheet["J7"].value
            sheet["F28"] = ((sheet["D28"].value)/60)*(sheet["E28"].value)
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Invoice")
    
    sheet = wb["Timesheet"]
    if app.getEntry("Notes 1") != "":
        sheet["I20"] = str(app.getEntry("Notes 1"))
    if app.getEntry("Notes 2") != "":
        sheet["I21"] = str(app.getEntry("Notes 2"))
    if app.getEntry("Notes 3") != "":
        sheet["I22"] = str(app.getEntry("Notes 3"))
    if app.getEntry("Notes 4") != "":
        sheet["I23"] = str(app.getEntry("Notes 4"))
    if app.getEntry("Notes 5") != "":
        sheet["I24"] = str(app.getEntry("Notes 5"))

    #Calendar time
    sheet = wb["Calendar"]
    dateNow = datetime.datetime.now()
    month = datetime.datetime.today().month

    global row, rowIndex, col, colIndex, day
    #If craeting the sheet, automatically fill out the days on the calendar.
    if created == 1:
        #Finds the current month.
        if month == 1:
            sheet["E7"] = "January"
        elif month == 2:
            sheet["E7"] = "February"
        elif month == 3:
            sheet["E7"] = "March"
        elif month == 4:
            sheet["E7"] = "April"
        elif month == 5:
            sheet["E7"] = "May"
        elif month == 6:
            sheet["E7"] = "June"
        elif month == 7:
            sheet["E7"] = "July"
        elif month == 8:
            sheet["E7"] = "August"
        elif month == 9:
            sheet["E7"] = "September"
        elif month == 10:
            sheet["E7"] = "October"
        elif month == 11:
            sheet["E7"] = "November"
        elif month == 12:
            sheet["E7"] = "December"
            
        #Finds when the first day (on which day; 0 is Mon, 6 is Sun.).
        first = dateNow.replace(day=1, hour=0, minute=0, second=0, microsecond=0).weekday()

        #Monday is cell 9 "first = 0"
        if first == 0:
            sheet["C9"] = 1
            colIndex = 1
        elif first == 1:
            sheet["D9"] = 1
            colIndex = 2
        elif first == 2:
            sheet["E9"] = 1
            colIndex = 3
        elif first == 3:
            sheet["F9"] = 1
            colIndex = 4
        elif first == 4:
            sheet["G9"] = 1
            colIndex = 5
        elif first == 5:
            sheet["H9"] = 1
            colIndex = 6
        elif first == 6:
            sheet["B9"] = 1
            colIndex = 0

        while True:
            colIndex += 1
            if month == 2 and day == 28:
                sheet[col[colIndex]+row[rowIndex]] = day
                break
            if (month == 4 or month == 6 or month == 9 or month == 11) and day == 30:
                sheet[col[colIndex]+row[rowIndex]] = day
                break
            if (month == 1 or month == 3 or month == 5 or month == 7 or month == 10 or month == 12) and day == 31:
                sheet[col[colIndex]+row[rowIndex]] = day
                break
            if colIndex < 7:
                sheet[col[colIndex]+row[rowIndex]] = day
                day += 1
            else:
                colIndex = 0
                if rowIndex < 4:
                    rowIndex += 1
                sheet[col[colIndex]+row[rowIndex]] = day
                day+=1
    
    app.setLabel("Month", str(sheet["E7"].value))

    #Get Info for today
    dateTodayWeek = datetime.datetime.today().weekday()
    dateToday = datetime.datetime.today().day

    if dateTodayWeek == 6:
        dateTodayWeek = -1

    for a in range(0,len(row)):
        if sheet[col[dateTodayWeek+1]+row[a]].value == dateToday:
            rowTime = int(row[a])
            app.setLabel("Get Info", str(sheet[col[dateTodayWeek+1]+str(rowTime+1)].value))
            break
    
    try:
        if created == 1:
            wb.save(og+"\\Invoices\\"+app.getEntry("New File Name")+".xlsx")
        elif created == 0:
            wb.save(path)
        
        app.showSubWindow("Calendar Today")
        app.hideSubWindow("Invoice")
        
    except PermissionError:
        app.errorBox("Error", "Permission denied when trying to write to the file. Try closing the file and try again.")
        app.clearAllEntries()
        app.showSubWindow("Invoice")
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Invoice")
    
def TodayOk(btn):
    try:
        dateTodayWeek = datetime.datetime.today().weekday()
        dateToday = datetime.datetime.today().day
        
        if dateTodayWeek == 6:
            dateTodayWeek = -1

        for a in range(0,len(row)):
            if sheet[col[dateTodayWeek+1]+row[a]].value == dateToday:
                rowTime = int(row[a])
                if app.getEntry("Today Note") != "":
                    sheet[col[dateTodayWeek+1]+str(rowTime+1)] = str(app.getEntry("Today Note"))
        
        app.hideSubWindow("Calendar Today")
        app.showSubWindow("Calendar")
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Calendar Today")

def CalendarOk(btn):
    try:
        infoWeekDay = datetime.datetime.now().replace(day=int(app.getOptionBox("Day"))).weekday()
        
        if infoWeekDay == 6:
            infoWeekDay = -1

        for a in range(0,len(row)):
            if sheet[col[infoWeekDay+1]+row[a]].value == int(app.getOptionBox("Day")):
                rowTime = int(row[a])
                app.setLabel("Existing", "Existing Info: "+str(sheet[col[infoWeekDay+1]+str(rowTime+1)].value))
                break 
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Calendar Today")
    
    try:
        if created == 1:
            wb.save(og+"\\Invoices\\"+app.getEntry("New File Name")+".xlsx")
        elif created == 0:
            wb.save(path)
        app.hideSubWindow("Calendar")
        app.showSubWindow("Calendar Enter")
    except PermissionError:
        app.errorBox("Error", "Permission denied when trying to write to the file. Try closing the file and try again.")
        app.clearAllEntries()
        app.showSubWindow("Calendar")
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Calendar")
    
def enter(btn):
    try:
        infoWeekDay = datetime.datetime.now().replace(day=int(app.getOptionBox("Day"))).weekday()
        
        if infoWeekDay == 6:
            infoWeekDay = -1

        for a in range(0,len(row)):
            if sheet[col[infoWeekDay+1]+row[a]].value == int(app.getOptionBox("Day")):
                rowTime = int(row[a])
                if app.getEntry("Add Note") != "":
                    sheet[col[infoWeekDay+1]+str(rowTime+1)] = str(app.getEntry("Add Note"))
                break 
        
        app.hideSubWindow("Calendar Enter")
        app.showSubWindow("Again")

    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Calendar Enter")

def againNo(btn):
    app.hideAllSubWindows()
    try:
        if created == 1:
            wb.save(og+"\\Invoices\\"+app.getEntry("New File Name")+".xlsx")
        elif created == 0:
            wb.save(path)
        app.hideSubWindow("Again")
        app.showSubWindow("Move")
    except PermissionError:
        app.errorBox("Error", "Permission denied when trying to write to the file. Try closing the file and try again.")
        app.clearAllEntries()
        app.showSubWindow("Again")
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Again")

def againYes(btn):
    try:
        app.hideSubWindow("Again")
        app.clearOptionBox("Day", callFunction = False)
        app.clearEntry("Add Note", callFunction = False)
        app.showSubWindow("Calendar")
    except:
        app.errorBox("Error", "An error has occured. Please try again or restart the program.")
        app.clearAllEntries()
        app.showSubWindow("Again")

def move(btn):
    if created == 1:
        try:
            shutil.move(og+"\\Invoices\\"+app.getEntry("New File Name")+".xlsx", app.getEntry("Move File"))
            app.hideSubWindow("Move")
            app.showSubWindow("End")
        except PermissionError:
            app.errorBox("Move Error", "Permission Denied. Please select another directory to move the file.\nThe file was not moved.")
            app.showSubWindow("Move")
        except FileNotFoundError:
            app.errorBox("Move Error", "Directory has either moved or does not exist. Please try again.")
            app.showSubWindow("Move")
        except:
            app.errorBox("Error", "An error has occured. Please try again or restart the program.")
            app.showSubWindow("Move")
    elif created == 0:
        try:
            shutil.move(path, app.getEntry("Move File"))
            app.showSubWindow("End")
        except PermissionError:
            app.errorBox("Move Error", "Permission Denied. Please select another directory to move the file.\nThe file was not moved.")
            app.showSubWindow("Move")
        except FileNotFoundError:
            app.errorBox("Move Error", "Directory has either moved or does not exist. Please try again.")
            app.showSubWindow("Move")
        except:
            app.errorBox("Error", "An error has occured. Please try again or restart the program.")
            app.showSubWindow("Move")
        

app = gui("Invoice Program for MSC Invoices", "600x400")
try:
    app.setIcon(icon)
except:
    pass
app.setBg("MistyRose")
app.createMenu("Help")
app.addMenuItem("Help", "About Me", about)
app.addMenuItem("Help", "Help", help)

app.startFrame("Intro", row = 0, column = 1)
app.addLabel("title", "Invoice Filling Program")
app.addLabel("author", "By Justin Chow")

app.setLabelBg("title", "FireBrick")
app.setLabelBg("author", "Royal blue")
app.setFont(size=12, family="Verdana")
app.stopFrame()

app.startFrame("Button 1", row = 1, column = 0)
app.addButton("Use a File", fileSend)
app.stopFrame()

app.startFrame("Button 2", row = 1, column = 2)
app.addButton("Create new file", createFile)
app.stopFrame()

app.startSubWindow("About Me!")
app.addLabel("This program is made by Justin Chow.\nProgram Version 1.0\nMade June 6th, 2020.\nThis program is to be used only for MSC Invoices.")
app.stopSubWindow()

app.startSubWindow("Create File")
app.addEntry("New File Name")
app.setEntryDefault("New File Name", "Enter a new file name")
app.addButton("Create File", checkName)
app.addNamedButton("Exit", "Create Exit", cancelOpen)
app.stopSubWindow()

app.startSubWindow("Use File")
app.addFileEntry("Invoice File")
app.setEntryDefault("Invoice File", "Add Invoice File here to update.")
app.addButton("Select", fileUse)
app.addNamedButton("Exit", "Use Exit", cancelOpen)
app.stopSubWindow()

app.startSubWindow("Invoice")
app.addLabel("Basic Info", row=0, column=0)
app.addLabel("Name: ", row=1, column=0)
app.addLabel("Wage: ", row=2, column=0)

app.addEntry("Name Input", row=1, column = 1)
app.setEntryTooltip("Name Input", "Enter your name")
app.addNumericEntry("Wage Input", row=2, column = 1)
app.setEntryTooltip("Wage Input", "Enter your wage")

app.addLabel("Lesson Type", row=3, column=0)
app.addLabel("PCS", row=4, column=0)
app.addLabel("CS", row=5, column=0)
app.addLabel("IN", row=6, column=0)
app.addLabel("PEP", row=7, column=0)
app.addLabel("AD", row=8, column=0)
app.addLabel("PW", row=9, column=0)
app.addLabel("ST", row=10, column=0)
app.addLabel("Off Ice", row=11, column=0)
app.addLabel("Off Ice Mins", row=12, column=0)
app.addLabel("Other", row=13, column=0)
app.addLabel("Other Mins", row=14, column=0)
app.addLabel("Notes Line 1", row=15, column=0)
app.addLabel("Notes Line 2", row=16, column=0)
app.addLabel("Notes Line 3", row=17, column=0)
app.addLabel("Notes Line 4", row=18, column=0)
app.addLabel("Notes Line 5", row=19, column=0)

app.addLabel("Number of Lessons", row=3, column=1)
app.addNumericEntry("PCS", row=4, column=1)
app.setEntryTooltip("PCS", "Enter the number of lessons")
app.addNumericEntry("CS", row=5, column=1)
app.setEntryTooltip("CS", "Enter the number of lessons")
app.addNumericEntry("IN", row=6, column=1)
app.setEntryTooltip("IN", "Enter the number of lessons")
app.addNumericEntry("PEP", row=7, column=1)
app.setEntryTooltip("PEP", "Enter the number of lessons")
app.addNumericEntry("AD", row=8, column=1)
app.setEntryTooltip("AD", "Enter the number of lessons")
app.addNumericEntry("PW", row=9, column=1)
app.setEntryTooltip("PW", "Enter the number of lessons")
app.addNumericEntry("ST", row=10, column=1)
app.setEntryTooltip("ST", "Enter the number of lessons")
app.addNumericEntry("Off Ice", row=11, column=1)
app.setEntryTooltip("Off Ice", "Enter the number of lessons")
app.addNumericEntry("Off Ice Mins", row=12, column=1)
app.setEntryTooltip("Off Ice Mins", "Enter the number of minutes spent on Off Ice lessons.\n Leave blank if no Off Ice lessons.")
app.addNumericEntry("Other", row=13, column=1)
app.setEntryTooltip("Other", "Enter the number of lessons")
app.addNumericEntry("Other Mins", row=14, column=1)
app.setEntryTooltip("Other Mins", "Enter the number of minutes spent on \"Other\" lessons.\nLeave blank if no \"Other\" lessons.")
app.addEntry("Notes 1", row=15, column = 1)
app.setEntryTooltip("Notes 1", "Enter the first line note")
app.addEntry("Notes 2", row=16, column = 1)
app.setEntryTooltip("Notes 2", "Enter the second line note")
app.addEntry("Notes 3", row=17, column = 1)
app.setEntryTooltip("Notes 3", "Enter the third line note")
app.addEntry("Notes 4", row=18, column = 1)
app.setEntryTooltip("Notes 4", "Enter the fourth line note")
app.addEntry("Notes 5", row=19, column = 1)
app.setEntryTooltip("Notes 5", "Enter the fifth line note")


app.addLabel("Current Values", row=0, column=2)
app.addLabel("Name Value", "Name: "+ str(sheet["J5"].value), row=1, column=2)
app.addLabel("Wage Value", "Wage: "+ str(sheet["J7"].value), row=2, column=2)
app.addLabel("PCS Value", "PCS: "+ str(sheet["C20"].value), row=4, column=2)
app.addLabel("CS Value", "CS: "+ str(sheet["C21"].value), row=5, column=2)
app.addLabel("IN Value", "IN: "+ str(sheet["C22"].value), row=6, column=2)
app.addLabel("PEP Value", "PEP: "+ str(sheet["C23"].value), row=7, column=2)
app.addLabel("AD Value", "AD: "+ str(sheet["C24"].value), row=8, column=2)
app.addLabel("PW Value", "PW: "+ str(sheet["C25"].value), row=9, column=2)
app.addLabel("ST Value", "ST: "+ str(sheet["C26"].value), row=10, column=2)
app.addLabel("Off-ice Value", "Off-ice: "+ str(sheet["C27"].value), row=11, column=2)
app.addLabel("Off-ice Min Value", "Off-ice Min: "+ str(sheet["D27"].value), row=12, column=2)
app.addLabel("Other Value", "Other: "+ str(sheet["C28"].value), row=13, column=2)
app.addLabel("Other Min Value", "Other Mins: "+ str(sheet["D28"].value), row=14, column=2)
app.addLabel("Notes 1 Value", "Notes 1: "+str(sheet["I20"].value), row=15, column=2)
app.addLabel("Notes 2 Value", "Notes 2: "+str(sheet["I21"].value), row=16, column=2)
app.addLabel("Notes 3 Value", "Notes 3: "+str(sheet["I22"].value), row=17, column=2)
app.addLabel("Notes 4 Value", "Notes 4: "+str(sheet["I23"].value), row=18, column=2)
app.addLabel("Notes 5 Value", "Notes 5: "+str(sheet["I24"].value), row=19, column=2)

app.addNamedButton("Ok","Invoice OK", accept, row=20, column=1)
app.addNamedButton("Abort", "Invoice Exit", abort, row=20, column=0)
app.stopSubWindow()

app.startSubWindow("Calendar Today")
app.addLabel("Update", "Update Info for today: ", row =0, column=0)
app.addEntry("Today Note", row=0, column = 1)
app.addLabel("Get Info", "", row=1, column=0)
app.addNamedButton("Abort", "Today Note Abort", abort, row=2, column=0)
app.addNamedButton("Ok", "Today Note Ok", TodayOk, row=2, column=1)
app.stopSubWindow()

app.startSubWindow("Calendar")
app.addLabel("Calendar Info", "Enter which date you wish to enter note for.")
app.addLabel("Month", str(sheet["E7"].value), row=1, column=0)
if month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10 or month == 12:
    app.addLabelOptionBox("Day", ["-Choose Day", "1", "2", "3", "4", "5", "6", "7", "8",
    "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27",
    "28", "29", "30", "31"], row=1, column=1)
elif month == 2 and leapyear == 0:
    app.addLabelOptionBox("Day", ["-Choose Day", "1", "2", "3", "4", "5", "6", "7", "8",
    "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27",
    "28"], row=1, column=1)
elif month == 2 and leapyear == 1:
    app.addLabelOptionBox("Day", ["-Choose Day", "1", "2", "3", "4", "5", "6", "7", "8",
    "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27",
    "28", "29"], row=1, column=1)
else:
    app.addLabelOptionBox("Day", ["-Choose Day", "1", "2", "3", "4", "5", "6", "7", "8",
    "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27",
    "28", "29", "30"], row=1, column=1)
app.addNamedButton("OK", "Calendar Ok", CalendarOk, row = 2, column = 2)
app.addNamedButton("Abort", "Calendar Abort", abort, row = 2, column = 0)
app.addNamedButton("Skip", "Calendar Skip", againNo, row = 2, column = 1)
app.stopSubWindow()

app.startSubWindow("Calendar Enter")
app.addLabel("Existing", "Existing: ", row=0)
app.addEntry("Add Note", row =1)
app.addNamedButton("Ok", "Any Ok", enter, row = 2, column =1)
app.addNamedButton("Abort", "Any Abort", abort, row = 2, column=0)
app.stopSubWindow()

app.startSubWindow("Again")
app.addLabel("Again", "Do you want to add another note for a different date?", row=0, column = 1)
app.addNamedButton("Abort", "Again Abort", abort, row=1, column=0)
app.addNamedButton("No", "Again No", againNo, row = 1, column = 1)
app.addNamedButton("Yes", "Again Yes", againYes, row = 1, column=2)
app.stopSubWindow()

app.startSubWindow("Move")
app.addLabel("Changes Saved! Select a location to move the invoice")
app.addDirectoryEntry("Move File")
app.addNamedButton("Move File", "Move", move)
app.addNamedButton("Close", "Move abort", abort)
app.stopSubWindow()

app.startSubWindow("End")
app.addLabel("You can now close this program.")
app.addNamedButton("Close", "End program", abort)
app.stopSubWindow()

app.go()